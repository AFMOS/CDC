import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
import logging
from bokeh.io import output_file, show
from bokeh.models import ColumnDataSource, DataTable, TableColumn, HTMLTemplateFormatter
from bokeh.layouts import column
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.formatting.rule import DataBarRule, IconSetRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.cell.cell import MergedCell

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Set the page configuration to wide mode
st.set_page_config(layout="wide")

# Main variable selector
main_variable_options = {"Total": "total", "Weight": "weight"}
main_variable = st.sidebar.radio(
    "Select Main Variable",
    options=list(main_variable_options.keys()),
    index=list(main_variable_options.keys()).index("Total")
)
main_variable = main_variable_options[main_variable]

# File uploader
uploaded_file = st.file_uploader("Choose a CSV or Excel file", type=['csv', 'xlsx'])

# Load the data
if uploaded_file is not None:
    if uploaded_file.name.endswith('.csv'):
        data = pd.read_csv(uploaded_file)
    elif uploaded_file.name.endswith('.xlsx'):
        data = pd.read_excel(uploaded_file)
else:
    try:
        data = pd.read_csv('sales_data.csv')
    except FileNotFoundError:
        st.error("The file 'sales_data.csv' was not found.")
        st.stop()
    except pd.errors.EmptyDataError:
        st.error("The file 'sales_data.csv' is empty.")
        st.stop()
    except pd.errors.ParserError:
        st.error("Error parsing 'sales_data.csv'. Please check the file format.")
        st.stop()

# Check for empty data
if data.empty:
    st.error("The loaded data is empty. Please check your data source.")
    st.stop()

# Define all possible months
all_months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Get the months present in the data
present_months = sorted(data['month'].unique(), key=lambda x: all_months.index(x))

# Use only the months present in the data for month_order
month_order = present_months

# Define quarter mapping
quarter_map = {
    'Q1': ["Jan", "Feb", "Mar"],
    'Q2': ["Apr", "May", "Jun"],
    'Q3': ["Jul", "Aug", "Sep"],
    'Q4': ["Oct", "Nov", "Dec"]
}

def apply_master_filter(df, search_term):
    if not search_term:
        return df, True

    search_term = search_term.lower()
    mask = pd.Series([False] * len(df))
    
    search_columns = ['customer_code', 'customer_name', 'customer_category', 'salesman', 
                      'item_code', 'item_description', 'item_category', 'month', 'area']
    
    for col in search_columns:
        if col in df.columns:
            mask |= df[col].astype(str).str.lower().str.contains(search_term, regex=True, na=False)
    
    filtered_df = df[mask]
    search_found = len(filtered_df) > 0
    return filtered_df, search_found

def update_filter_options(filtered_data):
    return {
        'area': ['None'] + list(filtered_data['area'].unique()),
        'month': ['None'] + list(filtered_data['month'].unique()),
        'quarter': ['None'] + list(quarter_map.keys()),
        'customer_category': ['None'] + list(filtered_data['customer_category'].unique()),
        'salesman': ['None'] + list(filtered_data['salesman'].unique()),
        'item_category': ['None'] + list(filtered_data['item_category'].unique()),
        'customer_name': ['None'] + list(filtered_data['customer_name'].unique()),
        'item_description': ['None'] + list(filtered_data['item_description'].unique())
    }

def generate_dashboard_title(search_term, selected_filters):
    title_parts = []
    
    if search_term:
        title_parts.append(f'<span style="color: maroon;">"{search_term}"</span>')
    
    for filter_name, filter_value in selected_filters.items():
        if filter_value and filter_value != "None":
            title_parts.append(f'<span style="color: maroon;">{filter_value}</span>')
    
    if title_parts:
        return f"{' - '.join(title_parts)} Breakdown"
    else:
        return "Sales Dashboard"

def create_heatmap_table(df):
    styled_df = df.copy()
    
    def change_indicator(value):
        if pd.isna(value):
            return ''
        elif value > 0:
            return '🟢'
        elif value < 0:
            return '🔴'
        else:
            return '⚪'

    styled_df['M.Change%'] = styled_df['M.Change%'].apply(lambda x: f"{change_indicator(x)} {x:+.2f}%" if pd.notnull(x) else '')
    
    num_colors = len(px.colors.sequential.Brwnyl)

    def style_numeric(value, column):
        if pd.isna(value):
            return ''
        if column in ['New Customers', 'Newly Listed Items']:
            max_val = df[column].max()
            color_index = int((value / max_val) * (num_colors - 1)) if max_val != 0 else 0
        elif column in ['Progress%', 'Customers Delt %', 'CTG Delt %']:
            color_index = int((value / 100) * (num_colors - 1))
        else:  # Sales column
            max_val = df['Sales'].max()
            color_index = int((value / max_val) * (num_colors - 1)) if max_val != 0 else 0
        
        color_index = max(0, min(color_index, num_colors - 1))
        color = px.colors.sequential.Brwnyl[color_index]
        return f'background-color: {color}; color: white;'

    def apply_style(x):
        col_name = x.name
        return [style_numeric(v, col_name) for v in x]

    styled = styled_df.style.apply(apply_style, axis=0, 
                                   subset=['Sales', 'New Customers', 'Newly Listed Items', 'Progress%', 'Customers Delt %', 'CTG Delt %'])
    
    return styled.format({
        'Sales': '{:,.0f}',
        'New Customers': '{:,.0f}',
        'Newly Listed Items': '{:.0f}',
        'M.Change%': '{}',
        'Progress%': '{:.2f}%',
        'Customers Delt %': '{:.2f}%',
        'CTG Delt %': '{:.2f}%'
    })

def update_dashboard(selected_area, selected_month, selected_quarter, 
                     selected_customer_category, selected_salesman, selected_item_category,
                     selected_customer_name, selected_item_description, main_variable, search_term):
    filtered_data, search_found = apply_master_filter(data, search_term)
    
    if not search_found or filtered_data.empty:
        return None, False

    # Apply filters
    if selected_area != "None":
        filtered_data = filtered_data[filtered_data['area'] == selected_area]
    
    if selected_month != "None" and selected_quarter == "None":
        filtered_data = filtered_data[filtered_data['month'] == selected_month]
    elif selected_quarter != "None" and selected_month == "None":
        months = quarter_map[selected_quarter]
        filtered_data = filtered_data[filtered_data['month'].isin(months)]
    elif selected_month != "None" and selected_quarter != "None":
        if selected_month not in quarter_map[selected_quarter]:
            st.warning(f"Selected month {selected_month} is not in the selected quarter {selected_quarter}. Applying only quarter filter.")
            months = quarter_map[selected_quarter]
            filtered_data = filtered_data[filtered_data['month'].isin(months)]
        else:
            filtered_data = filtered_data[filtered_data['month'] == selected_month]
    
    if selected_customer_category != "None":
        filtered_data = filtered_data[filtered_data['customer_category'] == selected_customer_category]
    if selected_salesman != "None":
        filtered_data = filtered_data[filtered_data['salesman'] == selected_salesman]
    if selected_item_category != "None":
        filtered_data = filtered_data[filtered_data['item_category'] == selected_item_category]
    if selected_customer_name != "None":
        filtered_data = filtered_data[filtered_data['customer_name'] == selected_customer_name]
    if selected_item_description != "None":
        filtered_data = filtered_data[filtered_data['item_description'] == selected_item_description]

    if filtered_data.empty:
        return None, False

    # Get unique months in the filtered data
    unique_months = sorted(filtered_data['month'].unique(), key=lambda x: all_months.index(x))

    # Calculate summary statistics
    total_sales = filtered_data[main_variable].sum()
    customer_count = filtered_data['customer_code'].nunique()
    total_weight = filtered_data['weight'].sum()
    
    total_customers = filtered_data['customer_code'].nunique()
    if total_customers > 0:
        Cash_count = filtered_data[filtered_data['payment_type'] == 'Cash']['customer_code'].nunique()
        Credit_count = filtered_data[filtered_data['payment_type'] == 'Credit']['customer_code'].nunique()
        Cash_percentage = Cash_count / total_customers
        Credit_percentage = Credit_count / total_customers
    else:
        Cash_count = Credit_count = 0
        Cash_percentage = Credit_percentage = 0

    # Create charts using Plotly
    # Sales by Area Pie Chart
    sales_by_area = filtered_data.groupby('area')[main_variable].sum().reset_index()
    fig_area = px.pie(sales_by_area, values=main_variable, names='area', title='Sales by Area')

    # Combined Time Graph (Monthly and Quarterly)
    sales_by_month = filtered_data.groupby('month')[main_variable].sum().reset_index()
    sales_by_month['month'] = pd.Categorical(sales_by_month['month'], categories=unique_months, ordered=True)
    sales_by_month = sales_by_month.sort_values('month')
    
    sales_by_quarter = filtered_data.copy()
    sales_by_quarter['quarter'] = sales_by_quarter['month'].apply(lambda x: next((q for q, months in quarter_map.items() if x in months), None))
    sales_by_quarter = sales_by_quarter.groupby('quarter')[main_variable].sum().reset_index()
    
    fig_time = go.Figure()

    fig_time.add_trace(go.Bar(
        x=sales_by_month['month'],
        y=sales_by_month[main_variable],
        name='Monthly Sales',
        marker_color='rgba(55, 83, 109, 0.7)',
        text=sales_by_month[main_variable].apply(lambda x: f'{x:,.0f}'),
        textposition='inside',
        textfont=dict(color='white'),
    ))

    if len(sales_by_month) > 1:
        sales_by_month['pct_change'] = sales_by_month[main_variable].pct_change() * 100
        for i, row in sales_by_month.iterrows():
            fig_time.add_annotation(
                x=row['month'],
                y=row[main_variable],
                text=f"{row['pct_change']:.1f}%" if not pd.isna(row['pct_change']) else "",
                showarrow=False,
                yshift=20,
                font=dict(size=10, color='rgba(55, 83, 109, 1)')
            )

    # Add quarterly data as lines with markers
    quarter_positions = {'Q1': 1, 'Q2': 4, 'Q3': 7, 'Q4': 10}
    valid_quarters = [q for q in sales_by_quarter['quarter'] if q in quarter_positions and quarter_positions[q]-1 < len(unique_months)]
    
    fig_time.add_trace(go.Scatter(
        x=[unique_months[quarter_positions[q]-1] for q in valid_quarters],
        y=sales_by_quarter[sales_by_quarter['quarter'].isin(valid_quarters)][main_variable],
        mode='lines+markers+text',
        name='Quarterly Sales',
        line=dict(color='rgba(255, 0, 0, 0.8)', width=3),
        marker=dict(size=12, symbol='star', color='rgba(255, 0, 0, 0.8)'),
        text=sales_by_quarter[sales_by_quarter['quarter'].isin(valid_quarters)][main_variable].apply(lambda x: f'{x:,.0f}'),
        textposition='top center',
        textfont=dict(color='rgba(255, 0, 0, 1)'),
    ))

    # Calculate and add percentage change for quarterly data
    if len(valid_quarters) > 1:
        sales_by_quarter['pct_change'] = sales_by_quarter[main_variable].pct_change() * 100
        for i, row in sales_by_quarter[sales_by_quarter['quarter'].isin(valid_quarters)].iterrows():
            if not pd.isna(row['pct_change']):
                fig_time.add_annotation(
                    x=unique_months[quarter_positions[row['quarter']]-1],
                    y=row[main_variable],
                    text=f"{row['pct_change']:.1f}%",
                    showarrow=False,
                    yshift=40,  # Increased yshift to avoid overlap with monthly annotations
                    font=dict(size=10, color='rgba(255, 0, 0, 0.8)')
                )

    fig_time.update_layout(
        title='Monthly and Quarterly Sales',
        xaxis_title='Month',
        yaxis_title=f'Sales ({main_variable.capitalize()})',
        legend_title='Period',
        hovermode="x unified",
        barmode='overlay',
        hoverlabel=dict(bgcolor="white", font_size=12),
        margin=dict(l=50, r=50, t=80, b=50),
    )
    fig_time.update_xaxes(tickangle=-45)

    # Sales by Salesman and Area Stacked Bar Chart
    sales_by_salesman_area = filtered_data.groupby(['salesman', 'area'])[main_variable].sum().reset_index()
    salesman_order = sales_by_salesman_area.groupby('salesman')[main_variable].sum().sort_values(ascending=False).index
    fig_salesman = px.bar(sales_by_salesman_area, 
                          x='salesman', 
                          y=main_variable, 
                          color='area',
                          title='Sales by Salesman and Area',
                          labels={main_variable: 'Total Sales', 'salesman': 'Salesman', 'area': 'Area'},
                          category_orders={'salesman': salesman_order},
                          text=main_variable)
    
    fig_salesman.update_traces(texttemplate='%{text:.0s}', textposition='inside')
    fig_salesman.update_layout(
        xaxis_title='',
        yaxis_title='',
        xaxis=dict(tickangle=-45),
        legend_title='Area',
        barmode='stack'
    )
    
    # Generate Item Category Heatmap
    pivot_table_item = filtered_data.pivot_table(index='item_category', columns='month', values=main_variable, aggfunc='sum')
    pivot_table_item = pivot_table_item.reindex(columns=unique_months)
    pivot_table_item['total'] = pivot_table_item.sum(axis=1)
    pivot_table_item = pivot_table_item.sort_values('total', ascending=True)
    pivot_table_item = pivot_table_item.drop('total', axis=1)
    
    if not pivot_table_item.empty:
        rounded_values_item = pivot_table_item.round(0).fillna('') 
        text_values_item = rounded_values_item.astype(str)
        text_values_item = text_values_item.replace('0.0', '', regex=False)
        
        fig_heatmap_item = go.Figure(data=go.Heatmap(
            z=pivot_table_item.values,
            x=pivot_table_item.columns,
            y=pivot_table_item.index,
            colorscale='Brwnyl',
            hoverongaps=False,
            text=text_values_item,
            texttemplate="%{text}",
            showscale=False
        ))
        
        fig_heatmap_item.update_layout(
            title='Item Category Heatmap',
            xaxis_title='',
            yaxis_title='',
            height=25 * len(pivot_table_item.index),
            xaxis=dict(side='top')
        )
    else:
        fig_heatmap_item = go.Figure()
        fig_heatmap_item.update_layout(title='No data available for Item Category Heatmap')

    # Generate Item Description Heatmap
    pivot_table_item_description = filtered_data.pivot_table(index='item_description', columns='month', values=main_variable, aggfunc='sum')
    pivot_table_item_description = pivot_table_item_description.reindex(columns=unique_months)
    pivot_table_item_description['total'] = pivot_table_item_description.sum(axis=1)
    pivot_table_item_description = pivot_table_item_description.sort_values('total', ascending=True)
    pivot_table_item_description = pivot_table_item_description.drop('total', axis=1)
    
    if not pivot_table_item_description.empty:
        rounded_values_item_description = pivot_table_item_description.round(0).fillna('')
        text_values_item_description = rounded_values_item_description.astype(str)
        text_values_item_description = text_values_item_description.replace('0.0', '', regex=False)
        
        fig_heatmap_item_description = go.Figure(data=go.Heatmap(
            z=pivot_table_item_description.values,
            x=pivot_table_item_description.columns,
            y=pivot_table_item_description.index,
            colorscale='Brwnyl',
            hoverongaps=False,
            text=text_values_item_description,
            texttemplate="%{text}",
            showscale=False
        ))
        
        fig_heatmap_item_description.update_layout(
            title='Item Description Heatmap',
            xaxis_title='',
            yaxis_title='',
            height=20 * len(pivot_table_item_description.index),
            xaxis=dict(side='top')
        )
    else:
        fig_heatmap_item_description = go.Figure()
        fig_heatmap_item_description.update_layout(title='No data available for Item Description Heatmap')

    # Generate Customer Heatmap
    pivot_table_customer = filtered_data.pivot_table(index='customer_name', columns='month', values=main_variable, aggfunc='sum')
    pivot_table_customer = pivot_table_customer.reindex(columns=unique_months)
    pivot_table_customer['total'] = pivot_table_customer.sum(axis=1)
    pivot_table_customer = pivot_table_customer.sort_values('total', ascending=True)
    pivot_table_customer = pivot_table_customer.drop('total', axis=1)
    
    if not pivot_table_customer.empty:
        rounded_values_customer = pivot_table_customer.round(0).fillna('')
        text_values_customer = rounded_values_customer.astype(str)
        text_values_customer = text_values_customer.replace('0.0', '', regex=False)
        
        fig_heatmap_customer = go.Figure(data=go.Heatmap(
            z=pivot_table_customer.values,
            x=pivot_table_customer.columns,
            y=pivot_table_customer.index,
            colorscale='Brwnyl',
            hoverongaps=False,
            text=text_values_customer,
            texttemplate="%{text}",
            showscale=False
        ))
        
        fig_heatmap_customer.update_layout(
            title='Customer Heatmap',
            xaxis_title='',
            yaxis_title='',
            height=20 * len(pivot_table_customer.index),
            xaxis=dict(side='top')
        )
    else:
        fig_heatmap_customer = go.Figure()
        fig_heatmap_customer.update_layout(title='No data available for Customer Heatmap')

    # Calculate monthly KPIs for Salesman tab
    monthly_kpis = filtered_data.groupby('month').agg({
        main_variable: 'sum',
        'customer_code': 'nunique',
        'item_category': 'nunique'
    }).reset_index()
    
    monthly_kpis['month'] = pd.Categorical(monthly_kpis['month'], categories=unique_months, ordered=True)
    monthly_kpis = monthly_kpis.sort_values('month')
    
    # Calculate new customers
    all_customers = set()
    new_customers = []
    new_customers_data = []
    for month in unique_months:
        month_data = filtered_data[filtered_data['month'] == month]
        month_customers = set(month_data['customer_code'])
        new_month_customers = month_customers - all_customers
        new_customers.append(len(new_month_customers))
        
        for customer in new_month_customers:
            customer_data = month_data[month_data['customer_code'] == customer]
            new_customers_data.append({
                'customer_code': customer,
                'month': month,
                'total': customer_data[main_variable].sum()
            })
        
        all_customers.update(month_customers)
    
    monthly_kpis['New Customer'] = new_customers

    # Calculate Newly Listed Items
    newly_listed = []
    all_combinations = set()
    newly_listed_items_data = []

    for month in unique_months:
        month_data = filtered_data[filtered_data['month'] == month]
        month_combinations = set(zip(month_data['customer_code'], month_data['item_code']))
        new_combinations = month_combinations - all_combinations
        
        valid_new_combinations = 0
        for customer_code, item_code in new_combinations:
            item_data = month_data[(month_data['customer_code'] == customer_code) & (month_data['item_code'] == item_code)]
            total = item_data[main_variable].sum()
            if total >= 1:
                valid_new_combinations += 1
                newly_listed_items_data.append({
                    'customer_code': customer_code,
                    'date': item_data['date'].min(),  # Use the earliest date for this combination
                    'item_code': item_code,
                    'total': total,
                    'salesman': item_data['salesman'].iloc[0]  # Add salesman information
                })
        
        newly_listed.append(valid_new_combinations)
        all_combinations.update(month_combinations)
    
    monthly_kpis['Newly Listed Items'] = newly_listed

    # Create DataFrames for export
    new_customers_df = pd.DataFrame(new_customers_data)
    newly_listed_items_df = pd.DataFrame(newly_listed_items_data)

    # Calculate additional KPIs
    if len(monthly_kpis) > 1:
        monthly_kpis['M.Change%'] = monthly_kpis[main_variable].pct_change() * 100
    else:
        monthly_kpis['M.Change%'] = 0
    
    total_sales = monthly_kpis[main_variable].sum()
    if total_sales > 0:
        monthly_kpis['Progress%'] = (monthly_kpis[main_variable].cumsum() / (total_sales * 1.5)) * 100
    else:
        monthly_kpis['Progress%'] = 0
    
    monthly_kpis['Customers Delt %'] = (monthly_kpis['customer_code'] / filtered_data['customer_code'].nunique()) * 100
    monthly_kpis['CTG Delt %'] = (monthly_kpis['item_category'] / filtered_data['item_category'].nunique()) * 100
    
    # Create a DataFrame for display
    display_df = pd.DataFrame({
        'Month': monthly_kpis['month'],
        'Sales': monthly_kpis[main_variable],
        'New Customers': monthly_kpis['New Customer'],
        'Newly Listed Items': monthly_kpis['Newly Listed Items'],
        'M.Change%': monthly_kpis['M.Change%'],
        'Progress%': monthly_kpis['Progress%'],
        'Customers Delt %': monthly_kpis['Customers Delt %'],
        'CTG Delt %': monthly_kpis['CTG Delt %']
    })
    
    display_df.set_index('Month', inplace=True)

    return (total_sales, customer_count, total_weight, Cash_count, Credit_count,
            Cash_percentage, Credit_percentage, fig_area, fig_time, fig_salesman, 
            fig_heatmap_item, fig_heatmap_item_description, fig_heatmap_customer, display_df,
            new_customers_df, newly_listed_items_df), True

def generate_excel(filtered_data, display_df):
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    # Define styles
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF", name="Dubai")
    header_style.fill = PatternFill(start_color="4f5966", end_color="4f5966", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center")
    header_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    data_style = NamedStyle(name="data_style")
    data_style.font = Font(name="Dubai")
    data_style.alignment = Alignment(horizontal="center", vertical="center")
    data_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    good_style = NamedStyle(name="good_style")
    good_style.font = Font(color="006100", name="Dubai")
    good_style.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    good_style.alignment = Alignment(horizontal="center", vertical="center")
    good_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    bad_style = NamedStyle(name="bad_style")
    bad_style.font = Font(color="9C0006", name="Dubai")
    bad_style.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    bad_style.alignment = Alignment(horizontal="center", vertical="center")
    bad_style.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Define the order of salesmen
    salesman_order = ["Amro Kendeel", "Salah Ali", "Mohammed Alwakeel", "Ibrahim Mohi", "Mohammed Jad", "Mohammed Saeed", "Mohammed Hasan"]
    
    # Convert all salesman names to strings and remove any potential NaN values
    all_salesmen = set(str(x) for x in filtered_data['salesman'].unique() if pd.notna(x))
    other_salesmen = sorted(all_salesmen - set(salesman_order))
    ordered_salesmen = salesman_order + other_salesmen

    for salesman in ordered_salesmen:
        if str(salesman) not in set(str(x) for x in filtered_data['salesman'].unique()):
            continue

        # Create a new worksheet for each salesman
        ws = wb.create_sheet(title=str(salesman))

        # Filter data for the current salesman
        salesman_data = filtered_data[filtered_data['salesman'].astype(str) == str(salesman)]
        
        # Recalculate KPIs for this salesman
        salesman_monthly_kpis = salesman_data.groupby('month').agg({
            'total': 'sum',
            'customer_code': 'nunique',
            'item_category': 'nunique'
        }).reset_index()
        
        salesman_monthly_kpis['month'] = pd.Categorical(salesman_monthly_kpis['month'], categories=month_order, ordered=True)
        salesman_monthly_kpis = salesman_monthly_kpis.sort_values('month')
        
        # Calculate new customers for this salesman
        all_customers = set()
        new_customers = []
        for month in salesman_monthly_kpis['month']:
            month_data = salesman_data[salesman_data['month'] == month]
            month_customers = set(month_data['customer_code'])
            new_month_customers = month_customers - all_customers
            new_customers.append(len(new_month_customers))
            all_customers.update(month_customers)
        
        salesman_monthly_kpis['New Customer'] = new_customers

        # Calculate Newly Listed Items for this salesman
        all_combinations = set()
        newly_listed = []
        for month in salesman_monthly_kpis['month']:
            month_data = salesman_data[salesman_data['month'] == month]
            month_combinations = set(zip(month_data['customer_code'], month_data['item_code']))
            new_combinations = month_combinations - all_combinations
            newly_listed.append(len(new_combinations))
            all_combinations.update(month_combinations)
        
        salesman_monthly_kpis['Newly Listed Items'] = newly_listed

        # Calculate additional KPIs for this salesman
        if len(salesman_monthly_kpis) > 1:
            salesman_monthly_kpis['M.Change%'] = salesman_monthly_kpis['total'].pct_change() * 100
        else:
            salesman_monthly_kpis['M.Change%'] = 0
        
        total_sales = salesman_monthly_kpis['total'].sum()
        if total_sales > 0:
            salesman_monthly_kpis['Progress%'] = (salesman_monthly_kpis['total'].cumsum() / (total_sales * 1.5)) * 100
        else:
            salesman_monthly_kpis['Progress%'] = 0
        
        salesman_monthly_kpis['Customers Delt %'] = (salesman_monthly_kpis['customer_code'] / salesman_data['customer_code'].nunique()) * 100
        salesman_monthly_kpis['CTG Delt %'] = (salesman_monthly_kpis['item_category'] / salesman_data['item_category'].nunique()) * 100

        # Create a DataFrame for display
        salesman_display_df = pd.DataFrame({
            'Month': salesman_monthly_kpis['month'],
            'Sales': salesman_monthly_kpis['total'],
            'New Customers': salesman_monthly_kpis['New Customer'],
            'Newly Listed Items': salesman_monthly_kpis['Newly Listed Items'],
            'M.Change%': salesman_monthly_kpis['M.Change%'],
            'Progress%': salesman_monthly_kpis['Progress%'],
            'Customers Delt %': salesman_monthly_kpis['Customers Delt %'],
            'CTG Delt %': salesman_monthly_kpis['CTG Delt %']
        })
        
        salesman_display_df.set_index('Month', inplace=True)
        
        # Write salesman name
        ws.cell(row=1, column=1, value=f"Salesman: {salesman}")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(salesman_display_df.columns) + 1)
        
        # Remove cell color and border for salesman name cell
        h2_cell = ws.cell(row=1, column=1)
        h2_cell.fill = PatternFill(fill_type=None)
        h2_cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

        # Write headers
        headers = ['Month'] + list(salesman_display_df.columns)
        if 'Progress%' in headers:
            headers.remove('Progress%')  # Remove Progress% column
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.style = header_style
        
        # Write data
        for idx, (month, row) in enumerate(salesman_display_df.iterrows(), start=3):
            ws.cell(row=idx, column=1, value=str(month))
            for col, header in enumerate(headers[1:], start=2):
                value = row[header]
                if pd.isna(value):
                    ws.cell(row=idx, column=col, value='')
                elif header in ['Sales', 'New Customers', 'Newly Listed Items']:
                    ws.cell(row=idx, column=col, value=int(value) if pd.notnull(value) else 0)
                elif header in ['M.Change%', 'Customers Delt %', 'CTG Delt %']:
                    ws.cell(row=idx, column=col, value=value / 100 if pd.notnull(value) else 0)
                else:
                    ws.cell(row=idx, column=col, value=str(value))

        # Apply styling and data bars
        data_rows = range(3, len(salesman_display_df) + 3)
        
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            
            # Get all values in the column
            column_values = [ws.cell(row=row, column=col).value for row in data_rows]
            # Filter out non-numeric values and convert to float
            numeric_values = []
            for value in column_values:
                try:
                    if isinstance(value, (int, float)):
                        numeric_values.append(float(value))
                    elif isinstance(value, str) and value.replace('.', '', 1).isdigit():
                        numeric_values.append(float(value))
                except (ValueError, TypeError):
                    pass
            
            if numeric_values:
                min_value = min(numeric_values)
                max_value = max(numeric_values)
            else:
                min_value = max_value = 0
            
            for row in data_rows:
                cell = ws.cell(row=row, column=col)
                cell.style = data_style
                
                if headers[col-1] == 'Sales':
                    cell.number_format = '#,##0'
                    rule = DataBarRule(start_type='num', start_value=min_value, end_type='num', end_value=max_value,
                                       color="FFA500", showValue=True, minLength=None, maxLength=None)
                    ws.conditional_formatting.add(f'{column_letter}{data_rows[0]}:{column_letter}{data_rows[-1]}', rule)
                elif headers[col-1] == 'New Customers':
                    cell.number_format = '#,##0'
                    rule = DataBarRule(start_type='num', start_value=min_value, end_type='num', end_value=max_value,
                                       color="4682B4", showValue=True, minLength=None, maxLength=None)
                    ws.conditional_formatting.add(f'{column_letter}{data_rows[0]}:{column_letter}{data_rows[-1]}', rule)
                elif headers[col-1] == 'Newly Listed Items':
                    cell.number_format = '#,##0'
                    rule = DataBarRule(start_type='num', start_value=min_value, end_type='num', end_value=max_value,
                                       color="FF69B4", showValue=True, minLength=None, maxLength=None)
                    ws.conditional_formatting.add(f'{column_letter}{data_rows[0]}:{column_letter}{data_rows[-1]}', rule)
                elif headers[col-1] == 'M.Change%':
                    cell.number_format = '0%'
                    value = cell.value
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.style = good_style
                        elif value < 0:
                            cell.style = bad_style
                    # Add traffic light icons
                    icon_set = IconSetRule('3TrafficLights1', 'percent', [0, 33, 67], showValue=True, percent=True, reverse=False)
                    ws.conditional_formatting.add(f'{column_letter}{data_rows[0]}:{column_letter}{data_rows[-1]}', icon_set)
                elif headers[col-1] == 'Customers Delt %':
                    cell.number_format = '0%'
                    rule = DataBarRule(start_type='num', start_value=min_value, end_type='num', end_value=max_value,
                                       color="800000", showValue=True, minLength=None, maxLength=None)
                    ws.conditional_formatting.add(f'{column_letter}{data_rows[0]}:{column_letter}{data_rows[-1]}', rule)
                elif headers[col-1] == 'CTG Delt %':
                    cell.number_format = '0%'
                    rule = DataBarRule(start_type='num', start_value=min_value, end_type='num', end_value=max_value,
                                       color="FA8072", showValue=True, minLength=None, maxLength=None)
                    ws.conditional_formatting.add(f'{column_letter}{data_rows[0]}:{column_letter}{data_rows[-1]}', rule)
                
                # Remove shading and border from the cell right of CTG Delt %
                if headers[col-1] == 'CTG Delt %':
                    next_cell = ws.cell(row=row, column=col+1)
                    next_cell.fill = PatternFill(fill_type=None)
                    next_cell.border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))

        # Adjust column widths
        for col in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

# Streamlit app
st.sidebar.header('Filters')

# Add master search filter
search_term = st.sidebar.text_input("Master Search Filter")

# Initial filter options
filter_options = update_filter_options(data)

# Apply filters and update options
filtered_data, _ = apply_master_filter(data, search_term)
filter_options = update_filter_options(filtered_data)

selected_area = st.sidebar.selectbox('Select Area', options=filter_options['area'])
selected_month = st.sidebar.selectbox('Select Month', options=filter_options['month'])
selected_quarter = st.sidebar.selectbox('Select Quarter', options=filter_options['quarter'])
selected_customer_category = st.sidebar.selectbox('Select Customer Category', options=filter_options['customer_category'])
selected_salesman = st.sidebar.selectbox('Select Salesman', options=filter_options['salesman'])
selected_item_category = st.sidebar.selectbox('Select Item Category', options=filter_options['item_category'])
selected_customer_name = st.sidebar.selectbox('Select Customer Name', options=filter_options['customer_name'])
selected_item_description = st.sidebar.selectbox('Select Item Description', options=filter_options['item_description'])

# Generate dynamic dashboard title
selected_filters = {
    'Area': selected_area,
    'Month': selected_month,
    'Quarter': selected_quarter,
    'Customer Category': selected_customer_category,
    'Salesman': selected_salesman,
    'Item Category': selected_item_category,
    'Customer Name': selected_customer_name,
    'Item Description': selected_item_description
}
dashboard_title = generate_dashboard_title(search_term, selected_filters)

# Display the dynamic dashboard title
st.markdown(f"""<h1 style="text-align: center;">📊 {dashboard_title}</h1>""", unsafe_allow_html=True)

# Update dashboard based on selections
try:
    dashboard_data, search_found = update_dashboard(
        selected_area, selected_month, selected_quarter, 
        selected_customer_category, selected_salesman, selected_item_category,
        selected_customer_name, selected_item_description,
        main_variable, search_term
    )
    
    if not search_found:
        st.warning("No data found matching the search criteria. Please adjust your filters or search term.")
    elif dashboard_data is None:
        st.warning("No data available for the selected filters. Please adjust your selections.")
    else:
        # Display dashboard components
        (total_sales, customer_count, total_weight, Cash_count, Credit_count,
         Cash_percentage, Credit_percentage, fig_area, fig_time, fig_salesman, 
         fig_heatmap_item, fig_heatmap_item_description, fig_heatmap_customer, display_df,
         new_customers_df, newly_listed_items_df) = dashboard_data

        # Layout with columns for summary statistics
        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown(f"""
            <div style="background-color: #f0f2f6; padding: 10px; border-radius: 5px; font-size: 12px; text-align: center;">
                <h4 style="margin: 0; font-size: 14px;">Total</h4>
                <h2 style="margin: 0; font-size: 16px;">
                {'SAR' if main_variable == 'total' else 'Kg'} {total_sales:,.0f}
                </h2>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            st.markdown(f"""
            <div style="background-color: #f0f2f6; padding: 10px; border-radius: 5px; font-size: 12px; text-align: center;">
                <h4 style="margin: 0; font-size: 14px;">Customer Count</h4>
                <h2 style="margin: 0; font-size: 16px;">{customer_count:,}</h2>
            </div>
            """, unsafe_allow_html=True)

        with col3:
            st.markdown(f"""
            <div style="background-color: #f0f2f6; padding: 10px; border-radius: 5px; font-size: 12px; text-align: center;">
                <h4 style="margin: 0; font-size: 14px;">Payment Type</h4>
                <h2 style="margin: 0; font-size: 16px;">Cash: {Cash_count:,} ({Cash_percentage:.1%})</h2>
                <h2 style="margin: 0; font-size: 16px;">Credit: {Credit_count:,} ({Credit_percentage:.1%})</h2>
            </div>
            """, unsafe_allow_html=True)

        # Create tabs for charts
        tab1, tab2, tab3, tab4 = st.tabs(["Sales Overview", "Time Graphs", "Heatmaps", "Salesman KPI"])

        with tab1:
            col1, col2 = st.columns(2)
            with col1:
                st.plotly_chart(fig_area, use_container_width=True)
            with col2:
                st.plotly_chart(fig_salesman, use_container_width=True)

        with tab2:
            st.plotly_chart(fig_time, use_container_width=True)

        with tab3:
            heatmap_option = st.selectbox(
                "Select Heatmap",
                ["Item Category", "Item Description", "Customer"]
            )
            
            if heatmap_option == "Item Category":
                st.plotly_chart(fig_heatmap_item, use_container_width=True)
            elif heatmap_option == "Item Description":
                st.plotly_chart(fig_heatmap_item_description, use_container_width=True)
            else:  # Customer
                st.plotly_chart(fig_heatmap_customer, use_container_width=True)

        with tab4:
            # Create two columns: one for KPI Heatmap, one for Export Options
            col1, col2 = st.columns([3, 1])
            
            with col1:
                # Display the heatmap-style table
                st.write("KPI Heatmap")
                if not display_df.empty:
                    try:
                        st.dataframe(create_heatmap_table(display_df), use_container_width=True)
                    except Exception as e:
                        st.error(f"Error creating heatmap table: {str(e)}")
                        logging.error(f"Heatmap table creation error: {str(e)}", exc_info=True)
                        st.write("Displaying raw data instead:")
                        st.dataframe(display_df)
                else:
                    st.write("No data available for KPI Heatmap")
            
            with col2:
                if not new_customers_df.empty:
                    csv = new_customers_df.to_csv(index=False)
                    st.download_button(
                        label="Export New Customers",
                        data=csv,
                        file_name="new_customers.csv",
                        mime="text/csv",
                    )
                else:
                    st.write("No new customers data available for export.")
                
                st.write("")  # Add some space between buttons
                
                if not newly_listed_items_df.empty:
                    valid_newly_listed_items_df = newly_listed_items_df[newly_listed_items_df['total'] >= 1]
                    csv = newly_listed_items_df.to_csv(index=False)
                    st.download_button(
                        label="Export Newly Listed Items",
                        data=csv,
                        file_name="newly_listed_items.csv",
                        mime="text/csv",
                    )
                else:
                    st.write("No newly listed items data available for export.")
                
                st.write("")  # Add some space before the new button
                
                # Add Export to Excel button
                if st.button("Export to Excel"):
                    excel_buffer = generate_excel(filtered_data, display_df)
                    st.download_button(
                        label="Download Excel",
                        data=excel_buffer,
                        file_name="salesman_reports.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

except Exception as e:
    st.error(f"An error occurred while updating the dashboard: {str(e)}")
    logging.error(f"Dashboard update error: {str(e)}", exc_info=True)
    st.stop()
